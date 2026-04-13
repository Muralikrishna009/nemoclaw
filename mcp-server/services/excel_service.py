"""
Excel generation service using openpyxl.
Produces styled .xlsx workbooks for financial, summary, invoice, and support reports.
Password encryption via msoffcrypto-tool.
"""

import os
import uuid
import io
import json
import urllib.request
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

try:
    import msoffcrypto
    _CRYPTO_AVAILABLE = True
except ImportError:
    _CRYPTO_AVAILABLE = False

from dummy_data.financial import (
    QUARTERLY_REVENUE, DEPARTMENTS, INVOICES, SUMMARY_STATS, SUPPORT_TICKETS,
)

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "outputs", "excels")
os.makedirs(OUTPUT_DIR, exist_ok=True)

_ADMIN_URL = os.getenv("ADMIN_PANEL_URL", "http://localhost:3000")

# ── Brand colours (hex strings for openpyxl) ──────────────────────────────────
C_PRIMARY   = "1A1A2E"
C_ACCENT    = "0F3460"
C_HIGHLIGHT = "E94560"
C_GREEN     = "27AE60"
C_LIGHT_BG  = "F8F9FA"
C_BORDER    = "DEE2E6"
C_WHITE     = "FFFFFF"
C_MUTED     = "888888"
C_YELLOW_BG = "FFF3CD"
C_GREEN_BG  = "D4EDDA"
C_RED_BG    = "F8D7DA"
C_GREY_BG   = "E2E3E5"


# ── Style helpers ──────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color=C_PRIMARY, size=10) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _border() -> Border:
    side = Side(style="thin", color=C_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center")


def _right() -> Alignment:
    return Alignment(horizontal="right", vertical="center")


def _header_row(ws, row: int, values: list, bg: str = C_PRIMARY):
    """Write a styled header row."""
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = _fill(bg)
        cell.font = _font(bold=True, color=C_WHITE, size=10)
        cell.alignment = _center()
        cell.border = _border()


def _data_row(ws, row: int, values: list, alt: bool = False,
              fills: dict = None, bold: bool = False, number_fmts: dict = None):
    """
    Write a data row.
    fills:       {col_index: hex_color}  — per-cell background override
    number_fmts: {col_index: format_str} — e.g. {2: '#,##0.00'}
    """
    bg = C_LIGHT_BG if alt else C_WHITE
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        override = (fills or {}).get(col)
        cell.fill = _fill(override if override else bg)
        cell.font = _font(bold=bold, color=C_PRIMARY)
        cell.alignment = _right() if isinstance(val, (int, float)) else _left()
        cell.border = _border()
        fmt = (number_fmts or {}).get(col)
        if fmt:
            cell.number_format = fmt


def _totals_row(ws, row: int, values: list):
    """Bold green-tinted totals row."""
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = _fill("E8F5E9")
        cell.font = _font(bold=True, color=C_PRIMARY)
        cell.alignment = _right() if isinstance(val, (int, float)) else _left()
        cell.border = _border()


def _set_col_widths(ws, widths: list):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _title_block(ws, title: str, subtitle: str):
    """Merge + style the top two rows as a title banner."""
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 20
    max_col = ws.max_column or 6
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)

    t = ws.cell(row=1, column=1, value=title)
    t.fill = _fill(C_PRIMARY)
    t.font = Font(bold=True, color=C_WHITE, size=16, name="Calibri")
    t.alignment = _center()

    s = ws.cell(row=2, column=1, value=subtitle)
    s.fill = _fill(C_ACCENT)
    s.font = Font(color="AAAACC", size=10, name="Calibri")
    s.alignment = _center()


def _kpi_block(ws, start_row: int, items: list):
    """
    items = [(label, value), ...]   — written as a 2-row label/value block.
    """
    ws.row_dimensions[start_row].height = 18
    ws.row_dimensions[start_row + 1].height = 24
    for col, (label, value) in enumerate(items, start=1):
        lc = ws.cell(row=start_row, column=col, value=label)
        lc.fill = _fill(C_ACCENT)
        lc.font = _font(bold=True, color=C_WHITE, size=9)
        lc.alignment = _center()
        lc.border = _border()

        vc = ws.cell(row=start_row + 1, column=col, value=value)
        vc.fill = _fill(C_LIGHT_BG)
        vc.font = _font(bold=True, color=C_PRIMARY, size=12)
        vc.alignment = _center()
        vc.border = _border()


# ── Password helper ────────────────────────────────────────────────────────────

def _get_file_password() -> str | None:
    try:
        req = urllib.request.Request(
            f"{_ADMIN_URL}/api/settings",
            headers={"Accept": "application/json"},
        )
        with urllib.request.urlopen(req, timeout=3) as resp:
            data = json.loads(resp.read())
        if data.get("pdf_password_enabled") == "true" and data.get("pdf_password"):
            return data["pdf_password"]
    except Exception:
        pass
    return None


def _save_workbook(wb: Workbook, filepath: str, password: str | None):
    """Save workbook, optionally encrypting with password."""
    if password and _CRYPTO_AVAILABLE:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        enc_buf = io.BytesIO()
        office_file = msoffcrypto.OfficeFile(buf)
        office_file.encrypt(password, enc_buf)
        enc_buf.seek(0)
        with open(filepath, "wb") as f:
            f.write(enc_buf.read())
    else:
        wb.save(filepath)


# ── Financial Excel ────────────────────────────────────────────────────────────

def generate_financial_excel(title: str, period: str, department: str | None = None,
                              password: str | None = None, **kwargs) -> str:
    filename = f"financial_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    wb = Workbook()
    generated = datetime.now().strftime("%B %d, %Y")

    # ── Sheet 1: Overview KPIs ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Overview"
    ws1.freeze_panes = "A5"

    # Title block needs max_col set first — use 5 cols
    for _ in range(5):
        ws1.append([])
    ws1.delete_rows(1, 5)

    _set_col_widths(ws1, [22, 22, 22, 22, 22])

    ws1.row_dimensions[1].height = 36
    ws1.row_dimensions[2].height = 20
    for col in range(1, 6):
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5) if col == 1 else None
        ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5) if col == 1 else None
        break

    t = ws1.cell(row=1, column=1, value=title or "Financial Report")
    t.fill = _fill(C_PRIMARY); t.font = Font(bold=True, color=C_WHITE, size=16, name="Calibri")
    t.alignment = _center()

    s = ws1.cell(row=2, column=1, value=f"Period: {period}  |  Generated: {generated}")
    s.fill = _fill(C_ACCENT); s.font = Font(color="AAAACC", size=10, name="Calibri")
    s.alignment = _center()

    ws1.row_dimensions[3].height = 8  # spacer

    _kpi_block(ws1, start_row=4, items=[
        ("TOTAL REVENUE",  f"${SUMMARY_STATS['total_revenue']:,.0f}"),
        ("TOTAL EXPENSES", f"${SUMMARY_STATS['total_expenses']:,.0f}"),
        ("NET PROFIT",     f"${SUMMARY_STATS['net_profit']:,.0f}"),
        ("PROFIT MARGIN",  f"{SUMMARY_STATS['profit_margin']}%"),
        ("YoY GROWTH",     f"{SUMMARY_STATS['yoy_growth']}%"),
    ])

    ws1.row_dimensions[6].height = 12  # spacer

    # Revenue breakdown table
    _header_row(ws1, 7, ["Month", "Revenue", "Expenses", "Profit", "Margin"])
    ws1.freeze_panes = "A8"

    quarter_data = QUARTERLY_REVENUE.get(period, list(QUARTERLY_REVENUE.values())[0])
    for i, row in enumerate(quarter_data):
        margin = round((row["profit"] / row["revenue"]) * 100, 1)
        _data_row(ws1, 8 + i,
                  [row["month"], row["revenue"], row["expenses"], row["profit"], f"{margin}%"],
                  alt=(i % 2 == 1),
                  number_fmts={2: '#,##0', 3: '#,##0', 4: '#,##0'})

    tot_r = sum(r["revenue"] for r in quarter_data)
    tot_e = sum(r["expenses"] for r in quarter_data)
    tot_p = sum(r["profit"] for r in quarter_data)
    _totals_row(ws1, 8 + len(quarter_data),
                ["TOTAL", tot_r, tot_e, tot_p,
                 f"{round(tot_p / tot_r * 100, 1)}%"])

    # ── Sheet 2: Department Budgets ────────────────────────────────────────────
    ws2 = wb.create_sheet("Departments")
    _set_col_widths(ws2, [20, 16, 16, 16, 12])
    ws2.freeze_panes = "A2"

    _header_row(ws2, 1, ["Department", "Budget", "Spent", "Remaining", "Headcount"])
    depts = {department: DEPARTMENTS[department]} \
        if department and department in DEPARTMENTS else DEPARTMENTS
    for i, (name, d) in enumerate(depts.items()):
        remaining = d["budget"] - d["spent"]
        pct = round(d["spent"] / d["budget"] * 100, 1)
        over = d["spent"] > d["budget"]
        _data_row(ws2, 2 + i,
                  [name, d["budget"], d["spent"], remaining, d["headcount"]],
                  alt=(i % 2 == 1),
                  fills={4: C_RED_BG if over else C_GREEN_BG},
                  number_fmts={2: '#,##0', 3: '#,##0', 4: '#,##0'})

    _save_workbook(wb, filepath, password)
    return filepath


# ── Summary Excel ──────────────────────────────────────────────────────────────

def generate_summary_excel(title: str, period: str, password: str | None = None, **kwargs) -> str:
    filename = f"summary_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    wb = Workbook()
    ws  = wb.active
    ws.title = "Executive Summary"
    generated = datetime.now().strftime("%B %d, %Y")
    _set_col_widths(ws, [16, 18, 18, 18, 14])

    # Title
    ws.merge_cells("A1:E1"); ws.merge_cells("A2:E2")
    ws.row_dimensions[1].height = 36; ws.row_dimensions[2].height = 20
    t = ws.cell(row=1, column=1, value=title or "Executive Summary")
    t.fill = _fill(C_PRIMARY); t.font = Font(bold=True, color=C_WHITE, size=16, name="Calibri")
    t.alignment = _center()
    s = ws.cell(row=2, column=1, value=f"Period: {period}  |  Generated: {generated}")
    s.fill = _fill(C_ACCENT); s.font = Font(color="AAAACC", size=10, name="Calibri")
    s.alignment = _center()

    # KPI row 1
    _kpi_block(ws, 4, [
        ("TOTAL REVENUE",  f"${SUMMARY_STATS['total_revenue']:,.0f}"),
        ("NET PROFIT",     f"${SUMMARY_STATS['net_profit']:,.0f}"),
        ("ACTIVE CLIENTS", str(SUMMARY_STATS["active_clients"])),
        ("NEW CLIENTS",    str(SUMMARY_STATS["new_clients"])),
        ("YoY GROWTH",     f"{SUMMARY_STATS['yoy_growth']}%"),
    ])

    # Quarterly performance table
    ws.row_dimensions[7].height = 10
    _header_row(ws, 8, ["Quarter", "Revenue", "Expenses", "Profit", "Margin"])
    ws.freeze_panes = "A9"

    for i, (q, months) in enumerate(QUARTERLY_REVENUE.items()):
        rev = sum(m["revenue"] for m in months)
        exp = sum(m["expenses"] for m in months)
        prf = sum(m["profit"] for m in months)
        _data_row(ws, 9 + i,
                  [q, rev, exp, prf, f"{round(prf / rev * 100, 1)}%"],
                  alt=(i % 2 == 1),
                  number_fmts={2: '#,##0', 3: '#,##0', 4: '#,##0'})

    _save_workbook(wb, filepath, password)
    return filepath


# ── Invoice Excel ──────────────────────────────────────────────────────────────

def generate_invoice_excel(title: str, period: str, password: str | None = None, **kwargs) -> str:
    filename = f"invoice_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    wb = Workbook()
    ws  = wb.active
    ws.title = "Invoice Ledger"
    generated = datetime.now().strftime("%B %d, %Y")
    _set_col_widths(ws, [14, 22, 16, 14, 14])

    # Title
    ws.merge_cells("A1:E1"); ws.merge_cells("A2:E2")
    ws.row_dimensions[1].height = 36; ws.row_dimensions[2].height = 20
    t = ws.cell(row=1, column=1, value=title or "Invoice Report")
    t.fill = _fill(C_PRIMARY); t.font = Font(bold=True, color=C_WHITE, size=16, name="Calibri")
    t.alignment = _center()
    s = ws.cell(row=2, column=1, value=f"Period: {period}  |  Generated: {generated}")
    s.fill = _fill(C_ACCENT); s.font = Font(color="AAAACC", size=10, name="Calibri")
    s.alignment = _center()

    total   = sum(i["amount"] for i in INVOICES)
    paid    = sum(i["amount"] for i in INVOICES if i["status"] == "Paid")
    pending = sum(i["amount"] for i in INVOICES if i["status"] == "Pending")
    overdue = sum(i["amount"] for i in INVOICES if i["status"] == "Overdue")

    _kpi_block(ws, 4, [
        ("TOTAL INVOICED", f"${total:,.0f}"),
        ("COLLECTED",      f"${paid:,.0f}"),
        ("PENDING",        f"${pending:,.0f}"),
        ("OVERDUE",        f"${overdue:,.0f}"),
        ("COUNT",          str(len(INVOICES))),
    ])

    ws.row_dimensions[7].height = 10
    _header_row(ws, 8, ["Invoice ID", "Client", "Amount", "Status", "Date"])
    ws.freeze_panes = "A9"

    status_fills = {"Paid": C_GREEN_BG, "Pending": C_YELLOW_BG, "Overdue": C_RED_BG}
    for i, inv in enumerate(INVOICES):
        _data_row(ws, 9 + i,
                  [inv["id"], inv["client"], inv["amount"], inv["status"], inv["date"]],
                  alt=(i % 2 == 1),
                  fills={4: status_fills.get(inv["status"], C_WHITE)},
                  number_fmts={3: '#,##0'})

    _totals_row(ws, 9 + len(INVOICES), ["", "TOTAL", total, "", ""])

    _save_workbook(wb, filepath, password)
    return filepath


# ── Support Excel ──────────────────────────────────────────────────────────────

def generate_support_excel(title: str, period: str, password: str | None = None, **kwargs) -> str:
    filename = f"support_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    wb = Workbook()
    st = SUPPORT_TICKETS
    generated = datetime.now().strftime("%B %d, %Y")

    # ── Sheet 1: Overview ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Overview"
    _set_col_widths(ws1, [24, 24, 24, 24])

    ws1.merge_cells("A1:D1"); ws1.merge_cells("A2:D2")
    ws1.row_dimensions[1].height = 36; ws1.row_dimensions[2].height = 20
    t = ws1.cell(row=1, column=1, value=title or "Support Tickets Report")
    t.fill = _fill(C_PRIMARY); t.font = Font(bold=True, color=C_WHITE, size=16, name="Calibri")
    t.alignment = _center()
    s = ws1.cell(row=2, column=1, value=f"Period: {period}  |  Generated: {generated}")
    s.fill = _fill(C_ACCENT); s.font = Font(color="AAAACC", size=10, name="Calibri")
    s.alignment = _center()

    _kpi_block(ws1, 4, [
        ("TOTAL TICKETS",  str(st["stats"]["total"])),
        ("OPEN",           str(st["stats"]["open"])),
        ("IN PROGRESS",    str(st["stats"]["in_progress"])),
        ("RESOLVED",       str(st["stats"]["resolved"])),
    ])
    _kpi_block(ws1, 7, [
        ("CLOSED",         str(st["stats"]["closed"])),
        ("AVG RESOLUTION", f"{st['stats']['avg_resolution_hours']}h"),
        ("SLA BREACH",     f"{st['stats']['sla_breach_rate']}%"),
        ("CSAT SCORE",     f"{st['stats']['csat_score']} / 5.0"),
    ])

    # ── Sheet 2: By Priority ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("By Priority")
    _set_col_widths(ws2, [16, 12, 12, 12, 18])
    ws2.freeze_panes = "A2"
    _header_row(ws2, 1, ["Priority", "Total", "Resolved", "Open", "Avg Resolution (h)"])

    priority_fills = {"Critical": C_RED_BG, "High": C_YELLOW_BG,
                      "Medium": C_LIGHT_BG, "Low": C_GREEN_BG}
    for i, r in enumerate(st["by_priority"]):
        _data_row(ws2, 2 + i,
                  [r["priority"], r["count"], r["resolved"],
                   r["count"] - r["resolved"], r["avg_hours"]],
                  fills={1: priority_fills.get(r["priority"], C_WHITE)})

    # ── Sheet 3: By Category ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("By Category")
    _set_col_widths(ws3, [22, 12, 12, 12, 18])
    ws3.freeze_panes = "A2"
    _header_row(ws3, 1, ["Category", "Total", "Resolved", "Open", "Resolution Rate"])

    for i, r in enumerate(st["by_category"]):
        rate = round(r["resolved"] / r["count"] * 100, 1) if r["count"] else 0
        _data_row(ws3, 2 + i,
                  [r["category"], r["count"], r["resolved"],
                   r["count"] - r["resolved"], f"{rate}%"],
                  alt=(i % 2 == 1))

    # ── Sheet 4: Recent Tickets ────────────────────────────────────────────────
    ws4 = wb.create_sheet("Recent Tickets")
    _set_col_widths(ws4, [12, 38, 12, 14, 14, 10])
    ws4.freeze_panes = "A2"
    _header_row(ws4, 1, ["Ticket ID", "Subject", "Priority", "Status", "Agent", "Hours"])

    status_fills = {"Resolved": C_GREEN_BG, "Closed": C_GREY_BG,
                    "In Progress": C_YELLOW_BG, "Open": C_RED_BG}
    for i, r in enumerate(st["recent"]):
        _data_row(ws4, 2 + i,
                  [r["id"], r["subject"], r["priority"], r["status"],
                   r["agent"], r["hours"] if r["hours"] else "—"],
                  fills={4: status_fills.get(r["status"], C_WHITE)})

    _save_workbook(wb, filepath, password)
    return filepath


# ── Dispatcher ─────────────────────────────────────────────────────────────────

EXCEL_GENERATORS = {
    "financial":       generate_financial_excel,
    "summary":         generate_summary_excel,
    "invoice":         generate_invoice_excel,
    "support":         generate_support_excel,
    "support_tickets": generate_support_excel,
}


def generate_excel(report_type: str, title: str, params: dict) -> str:
    period     = params.get("period", "Q1 2025")
    department = params.get("department")
    password   = _get_file_password()
    generator  = EXCEL_GENERATORS.get(report_type)
    if generator is None:
        return f"ERROR: Unknown report_type '{report_type}'. Supported: {', '.join(EXCEL_GENERATORS)}"
    return generator(title=title, period=period, department=department, password=password)
